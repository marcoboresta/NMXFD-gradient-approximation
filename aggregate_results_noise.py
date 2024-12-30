"""
Aggregate and Summarize Gradient Approximation Results (Noise Version)

This script processes Excel files containing gradient approximation results and
computes the median log errors for various methods. It organizes the data by
bucket, sigma, and N, and updates a template Excel file to summarize results.

Author: Marco Boresta
Date: Refactored: [27/12/2024]
"""

import os
import glob
import pandas as pd
import numpy as np
from openpyxl import load_workbook


def get_files_in_bucket(folder, bucket_number):
    """
    Retrieve the list of Excel files for a specific bucket.

    Args:
        folder (str): Path to the folder containing the Excel files.
        bucket_number (str): The bucket number (as a string).

    Returns:
        list: File paths matching the bucket.
    """
    files = glob.glob(os.path.join(folder, "*.xlsx"))
    return [file for file in files if f"bucket{bucket_number}" in file]


def process_file(file):
    """
    Process an individual Excel file and compute median log errors for each method.

    Args:
        file (str): Path to the Excel file.

    Returns:
        dict: Median log values and associated metadata for each method.
    """
    df = pd.read_excel(file)

    # Extract N and sigma values from the filename
    # Example filename: "bucket4_N13_sigma-0.001_noise-0.001.xlsx"
    # We split on "_N" then take everything until the next underscore for N,
    # and split on "sigma-" then take everything until the next underscore for sigma.
    N_part = file.split('_N')[1].split('_')[0]      # e.g. "13" or "5"
    sigma_part = file.split('sigma-')[1].split('_')[0]  # e.g. "0.001"

    # Convert them to strings for the final dictionary
    N = str(N_part)
    sigma = str(sigma_part)

    # The original script expects columns: NMXFD, FFD, CFD, ACFD, GSG, cGSG
    methods = ['NMXFD', 'FFD', 'CFD', 'ACFD', 'GSG', 'cGSG']

    # Compute log values for gradient approximation methods
    for method in methods:
        df[f'Log{method}'] = np.log10(df[method] + 1e-12)

    # Compute medians
    bucket_id = file.split("bucket")[1].split("_")[0]  # e.g. "4" from "bucket4_N..."
    medians = {
        "Bucket": bucket_id,
        "N": N,
        "Sigma": sigma,
    }
    medians.update({
        f"Median Log{method}": df[f'Log{method}'].median()
        for method in methods
    })

    return medians


def update_template(template_file, input_folder, output_folder):
    """
    Update a template Excel file with median values and save new summaries.

    Args:
        template_file (str): Path to the template Excel file.
        input_folder (str): Directory containing input Excel files.
        output_folder (str): Directory to save the updated files.
    """
    os.makedirs(output_folder, exist_ok=True)

    # Organize files by bucket
    files_by_bucket = {}
    for file in os.listdir(input_folder):
        if file.endswith('.xlsx') and file.startswith('bucket'):
            bucket = file.split('_')[0].replace('bucket', '')
            files_by_bucket.setdefault(bucket, []).append(file)

    # We may need an N-mapping for 5, 9, 13, similar to the original logic
    # The original snippet had: { '9': '8n', '5': '4n', '3': '2n' }
    # We'll add '13': '12n' so it doesn't fail
    indiceN_map = {
        '5':  '4n',
        '9':  '8n',
        '13': '12n'
    }

    # Columns to update
    approximations = {
        "Median LogFFD":  "FFD",
        "Median LogCFD":  "CFD",
        "Median LogGSG":  "GSG",
        "Median LogcGSG": "cGSG",
        "Median LogNMXFD": "NMXFD",
        "Median LogACFD":  "ACFD"
    }

    # Process each bucket
    for bucket, file_list in files_by_bucket.items():
        # Load the template workbook for each bucket (or load just once, if desired)
        workbook = load_workbook(template_file)
        sheet = workbook.active  # Assuming the template has a single sheet

        for file in file_list:
            file_path = os.path.join(input_folder, file)

            # Extract N and sigma
            # e.g. "bucket4_N13_sigma-0.001_noise-0.001.xlsx"
            N_part = file.split('_N')[1].split('_')[0]
            sigma_part = file.split('sigma-')[1].split('_')[0]

            # Validate we can handle N
            if N_part not in indiceN_map:
                raise ValueError(f"Unexpected N value: {N_part}")
            indiceN = indiceN_map[N_part]

            # Read the file into a DataFrame
            df = pd.read_excel(file_path)
            # df.rename(columns = {"GS": "GSG", "GSC": "cGSG", "FCD": "CFD"}, inplace=True)

            # For each approximation, compute median log10
            for column_name, approx_name in approximations.items():
                # Make sure the column is present
                if approx_name in df.columns:
                    median_value = np.log10(df[approx_name] + 1e-12).median()


                    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                        if row[0].value == approx_name:
                            if approx_name in ["FFD", "CFD"]:
                                # If method is FFD, handle specially
                                for col in range(3, sheet.max_column + 1):
                                    if sheet.cell(row=1, column=col).value == sigma_part:
                                        sheet.cell(row=row[0].row, column=col, value=median_value)
                            elif row[1].value == indiceN:
                                for col in range(3, sheet.max_column + 1):
                                    if sheet.cell(row=1, column=col).value == sigma_part:
                                        sheet.cell(row=row[0].row, column=col, value=median_value)

        # Save the updated Excel after processing all files in this bucket
        output_file = os.path.join(output_folder, f"bucket_{bucket}_median_noise.xlsx")
        workbook.save(output_file)
        print(f"Updated file saved: {output_file}")


if __name__ == "__main__":
    # Example usage:
    template_file = 'template_summary_file_noise.xlsx'  # Updated name
    input_folder = "Results/noise/"                   # Folder containing your new .xlsx results
    output_folder = "Results/summaries"        # Folder to store final aggregated templates

    update_template(template_file, input_folder, output_folder)
