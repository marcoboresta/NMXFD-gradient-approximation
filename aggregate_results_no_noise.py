"""
Aggregate and summarize results from individual Excel files.

This script computes the log and median of the errors for each method across 
multiple Excel files (grouped by bucket, sigma, and N). It generates a final
summary table with the medians for each method.

Author: Marco Boresta
Date: Refactored [27/12/2024]
"""

import os
import glob
import pandas as pd
import numpy as np
from openpyxl import load_workbook


def get_files_in_bucket(folder, bucket_number):
    """
    Retrieve the list of files for a specific bucket.

    Args:
        folder (str): Path to the folder containing the Excel files.
        bucket_number (str): The bucket number as a string.

    Returns:
        list: List of file paths matching the bucket.
    """
    files = glob.glob(os.path.join(folder, "*.xlsx"))
    return [file for file in files if f"bucket{bucket_number}" in file]


def process_file(file):
    """
    Process a single file to compute the median log values for each method.

    Args:
        file (str): Path to the Excel file.

    Returns:
        dict: Median log values for each method and associated metadata.
    """
    df = pd.read_excel(file)

    # Extract N and sigma from the filename
    N = file.split('_N')[1][0]
    sigma = file.split('sigma-')[1].split('_')[0]

    # Compute the log values for each method
    df['LogNMXFD'] = np.log10(df['NMXFD'] + 1e-12)
    df['LogFFD'] = np.log10(df['FFD'] + 1e-12)
    df['LogCFD'] = np.log10(df['CFD'] + 1e-12)
    df['LogACFD'] = np.log10(df['ACFD'] + 1e-12)
    df['LogGSG'] = np.log10(df['GSG'] + 1e-12)
    df['LogcGSG'] = np.log10(df['cGSG'] + 1e-12)

    # Compute the median log values
    medians = {
        "Bucket": file.split("bucket")[1].split("_")[0],  # Extract bucket number
        "N": N,
        "Sigma": sigma,
        "Median LogNMXFD": df['LogNMXFD'].median(),
        "Median LogFFD": df['LogFFD'].median(),
        "Median LogCFD": df['LogCFD'].median(),
        "Median LogACFD": df['LogACFD'].median(),
        "Median LogGSG": df['LogGSG'].median(),
        "Median LogcGSG": df['LogcGSG'].median(),
    }
    return medians


def update_template(template_file, input_folder, output_folder):
    """
    Update the template Excel file with median values from the input files,
    creating one output file per bucket.

    Args:
        template_file (str): Path to the template Excel file.
        input_folder (str): Path to the folder containing input Excel files.
        output_folder (str): Path to save the updated Excel files.
    """
    os.makedirs(output_folder, exist_ok=True)

    # Group files by bucket
    files_by_bucket = {}
    for file in os.listdir(input_folder):
        if file.endswith('.xlsx') and file.startswith('bucket'):
            bucket = file.split('_')[0].replace('bucket', '')
            if bucket not in files_by_bucket:
                files_by_bucket[bucket] = []
            files_by_bucket[bucket].append(file)

    # Process each bucket
    for bucket, file_list in files_by_bucket.items():
        # Load the template for the current bucket
        workbook = load_workbook(template_file)
        sheet = workbook.active  # Assumes the template has a single sheet

        # Define mappings for columns and approximations
        approximations = {
            "Median LogFFD": "FFD",
            "Median LogCFD": "CFD",
            "Median LogGSG": "GSG",
            "Median LogcGSG": "cGSG",
            "Median LogNMXFD": "NMXFD",
            "Median LogACFD": "ACFD"
        }

        # Process all files for the current bucket
        for file in file_list:
            # Extract N and sigma from the filename
            N = file.split('_N')[1][0]
            sigma = file.split('sigma-')[1].split('.xlsx')[0]

            # Map N to indiceN
            if N == '9':
                indiceN = '8n'
            elif N == '5':
                indiceN = '4n'
            elif N == '3':
                indiceN = '2n'
            else:
                raise ValueError(f"Unexpected N value: {N}")

            # Read the input Excel file
            file_path = os.path.join(input_folder, file)
            df = pd.read_excel(file_path)

            # Iterate over approximations and update the template
            for column_name, approx_name in approximations.items():
                if approx_name in df.columns:
                    # print(approx_name)
                    median_value = np.log10(df[approx_name]).median()

                    # Find the row and column in the template
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                    if row[0].value == approx_name:
                        # Check if it's FFD (independent of N)
                        if approx_name == "FFD":
                            # Find the correct sigma column
                            for col in range(3, sheet.max_column + 1):  # Sigma columns start from column 3
                                if sheet.cell(row=1, column=col).value == sigma:
                                    # Update the cell with the median value
                                    sheet.cell(row=row[0].row, column=col, value=median_value)
                        # Otherwise, check for N and write the value
                        elif row[1].value == indiceN:
                            # Find the correct sigma column
                            for col in range(3, sheet.max_column + 1):  # Sigma columns start from column 3
                                if sheet.cell(row=1, column=col).value == sigma:
                                    # Update the cell with the median value
                                    sheet.cell(row=row[0].row, column=col, value=median_value)

        # Save the updated file for the current bucket
        output_file = os.path.join(output_folder, f"bucket_{bucket}_median_no_noise.xlsx")
        workbook.save(output_file)
        print(f"Updated file saved: {output_file}")


if __name__ == "__main__":
    # Input and output folder paths
    template_file = 'template_summary_file_no_noise.xlsx'
    input_folder = "Results/no_noise"
    output_folder = "Results/summaries"
    update_template(template_file, input_folder, output_folder)

